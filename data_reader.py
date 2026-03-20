"""
폴더 스캔 기반 schema-agnostic 데이터 리더.

WATCH_DIR 안의 CSV/Excel 파일을 자동 스캔하고,
각 파일의 1행(헤더)만 읽어 컬럼명으로 카테고리(재고/생산/재무 등)를 자동 분류합니다.
"""

import csv
import os
import glob
import time
import re
import unicodedata
from datetime import date, datetime, timedelta
import numpy as np
from sentence_transformers import SentenceTransformer

from app.services.embedding_cache import load_row_embedding_cache

# ── 설정 ──────────────────────────────────────────────────────────────────────
# 여기를 원하는 폴더 경로로 변경하세요.
# 환경변수 DATA_DIR이 있으면 우선 사용합니다.
WATCH_DIR = os.environ.get(
    "DATA_DIR",
    os.path.join(os.path.dirname(__file__), "data")
)

SCAN_INTERVAL = 30  # 파일 목록을 다시 스캔할 간격(초). 0이면 매 요청마다 스캔.
CACHE_MAX_ENTRIES = 32  # 파일 내용 캐시 최대 개수. mtime 변경 시 자동 무효화.
# ─────────────────────────────────────────────────────────────────────────────

LABELS = ["재고", "생산", "재무", "규율", "기타"]

# 수치 조건(단가 N만원 이하 등)이 있을 때 행 텍스트 매칭에서 빼는 말(행에 안 나오는 조사형·범주어)
_NUMERIC_QUERY_SKIP_TOKENS = frozenset(
    {
        "제품",
        "품목",
        "상품",
        "아이템",
        "물건",
        "자재",
        "product",
        "item",
    }
)

# 카테고리별 헤더(1행) 컬럼 키워드 — 파일명 없이 1행만 읽어서 분류
HEADER_KEYWORDS: dict[str, list[str]] = {
    "재고": ["품목", "재고", "수량", "창고", "입고", "출고", "자재", "안전재고", "발주", "단가", "inventory", "stock"],
    "생산": ["생산", "라인", "공정", "가동", "설비", "불량", "수율", "작업일자", "production", "line"],
    "재무": ["매출", "비용", "예산", "손익", "미수금", "계정과목", "전표", "finance", "revenue", "expense"],
    "규율": ["규정", "근태", "출근", "퇴근", "연차", "휴가", "policy", "규율", "적용대상"],
    "기타": ["공지", "분류", "항목", "담당", "게시판", "포털", "misc", "notice"],
}

# 데이터 조회 대상에서 제외할 파일
IGNORE_FILENAMES = {
    "random_questions_1000.csv",
    "random_questions_1000.jsonl",
    "clarified_training_dataset.jsonl",
}

# 카테고리 대표 설명 (컬럼명 임베딩 2차 매칭용)
CATEGORY_DESCRIPTIONS: dict[str, str] = {
    "재고": "품목 수량 재고 입출고 창고 자재 부품 안전재고",
    "생산": "생산 라인 공정 설비 가동률 달성률 불량 작업",
    "재무": "매출 비용 예산 급여 보너스 이익 정산 회계",
    "규율": "사내 규정 근태 연차 휴가 보안 복무 신청 절차 정책",
    "기타": "사내 공지 게시판 포털 사용법 조직도 양식 시스템 안내",
}

# 로컬/서버 GPU에서 Qwen3-Embedding-8B 사용 시: EMBEDDING_MODEL=Qwen/Qwen3-Embedding-8B (VRAM 16GB+ 권장)
_EMBEDDING_MODEL_ID = os.environ.get("EMBEDDING_MODEL", "nlpai-lab/KURE-v1")
_IS_QWEN3_EMBEDDING = "Qwen3" in _EMBEDDING_MODEL_ID or "Qwen3-Embedding" in _EMBEDDING_MODEL_ID

# 사전 계산된 행 임베딩 캐시 사용 (scripts/build_embedding_cache.py). 끄려면 USE_ROW_EMBEDDING_CACHE=0
_USE_ROW_EMB_CACHE = os.environ.get("USE_ROW_EMBEDDING_CACHE", "1").strip().lower() not in (
    "0",
    "false",
    "no",
    "off",
)
# 키워드(토큰)로 행을 먼저 거른 뒤 임베딩. 기본은 끔 → 수치/입출고 필터 후 바로 임베딩 검색.
_USE_KEYWORD_ROW_FILTER = os.environ.get("USE_KEYWORD_ROW_FILTER", "0").strip().lower() in (
    "1",
    "true",
    "yes",
    "on",
)


def _parse_row_embed_sim_threshold() -> float | None:
    """
    행 임베딩 필터.
    - None: 임계값 없음 → 남은 행 전부, 코사인 유사도 높은 순 정렬
    - float: 해당 값 이상인 행만 (예: 0.5). 하나도 없으면 기존처럼 필터만 적용된 전체 rows
    """
    raw = (os.environ.get("ROW_EMBED_SIM_THRESHOLD") or "all").strip().lower()
    if raw in ("all", "none", "any", "off"):
        return None
    try:
        v = float(raw)
        return None if v <= 0 else v
    except ValueError:
        return None


def _parse_row_embed_top_k() -> int:
    try:
        return max(0, int(os.environ.get("ROW_EMBED_TOP_K", "0") or 0))
    except ValueError:
        return 0


_ROW_EMBED_SIM_THRESHOLD: float | None = _parse_row_embed_sim_threshold()
_ROW_EMBED_TOP_K: int = _parse_row_embed_top_k()
# (절대경로, mtime, text_cols 튜플, 모델 id) -> 전체 행 임베딩 행렬
_FULL_ROW_EMB_MEM: dict[tuple[str, float, tuple[str, ...], str], np.ndarray] = {}

def _load_embedding_model():
    kwargs = {}
    if _IS_QWEN3_EMBEDDING:
        kwargs["model_kwargs"] = {"device_map": "auto"}
        kwargs["tokenizer_kwargs"] = {"padding_side": "left"}
        if os.environ.get("USE_FLASH_ATTENTION", "").lower() in ("1", "true", "yes"):
            kwargs["model_kwargs"]["attn_implementation"] = "flash_attention_2"
    return SentenceTransformer(_EMBEDDING_MODEL_ID, **kwargs)

_model = _load_embedding_model()

def _encode_query(text_or_list, **kw):
    """질문(쿼리)용 인코딩. Qwen3-Embedding 시 prompt_name='query' 사용."""
    if _IS_QWEN3_EMBEDDING:
        kw["prompt_name"] = "query"
    kw.setdefault("convert_to_numpy", True)
    return _model.encode(text_or_list, **kw)

def _encode_docs(text_or_list, **kw):
    """문서/행/컬럼용 인코딩 (Qwen3는 프롬프트 없음)."""
    kw.setdefault("convert_to_numpy", True)
    return _model.encode(text_or_list, **kw)

_cat_desc_embs: dict[str, np.ndarray] = {
    label: _encode_docs(desc)
    for label, desc in CATEGORY_DESCRIPTIONS.items()
}

# 파일 인덱스 캐시: { label: filepath }
_file_index: dict[str, str] = {}
_last_scan: float = 0.0

# 파일 내용 캐시: path -> (mtime, rows). mtime이 바뀌면 재읽기.
_file_cache: dict[str, tuple[float, list[dict]]] = {}


# ── 유틸 ──────────────────────────────────────────────────────────────────────

def _make_headers_unique(headers: list[str]) -> list[str]:
    """컬럼명이 중복되면 뒤쪽에 _2, _3 ... 을 붙여 고유하게 만듭니다."""
    result: list[str] = []
    for h in headers:
        base = h
        name = h
        suffix = 2
        while name in result:
            name = f"{base}_{suffix}"
            suffix += 1
        result.append(name)
    return result


def _normalize_headers(first_row: list, second_row: list | None = None) -> list[str]:
    """
    1행을 헤더로 쓰고, 빈 칸은 2행 값으로 보정합니다.
    col0, col10 같은 무의미한 이름을 줄이기 위함.
    컬럼명 중복 시 _2, _3 ... 으로 고유화합니다.
    """
    ncols = len(first_row)
    raw: list[str] = []
    for i in range(ncols):
        h = first_row[i] if i < len(first_row) else None
        cell = str(h).strip() if (h is not None and str(h).strip()) else ""
        is_generic = not cell or (cell.startswith("col") and len(cell) > 3 and cell[3:].isdigit())
        if is_generic and second_row and i < len(second_row):
            s = second_row[i]
            s = str(s).strip() if s is not None else ""
            if s and len(s) <= 50 and not (s.replace(".", "").replace("-", "").replace("%", "").isdigit()):
                cell = s
            else:
                cell = cell if cell else f"col{i}"
        else:
            cell = cell if cell else f"col{i}"
        raw.append(cell)
    return _make_headers_unique(raw)


# ─────────────────────────────────────────────────────────────────────────────

def _cosine_sim(a: np.ndarray, b: np.ndarray) -> float:
    denom = np.linalg.norm(a) * np.linalg.norm(b)
    return float(np.dot(a, b) / denom) if denom > 1e-10 else 0.0


def _embed_rank_numeric_columns(candidates: list[str], query: str) -> list[str]:
    """
    column_hint가 없고 금액/수량 후보 열이 여럿일 때, 질문과 컬럼명 임베딩 유사도 순으로 정렬합니다.
    (행 값이 아니라 헤더 이름과 질문의 의미 정렬 — 수식 비교 전 열 선택용)
    """
    if len(candidates) <= 1 or not (query or "").strip():
        return candidates
    try:
        qe = _encode_query(query.strip(), show_progress_bar=False)
        qe = np.asarray(qe)
        if qe.ndim == 2:
            qe = qe.ravel()
        doc_embs = _encode_docs(candidates, show_progress_bar=False)
        doc_embs = np.asarray(doc_embs)
        if doc_embs.ndim == 1:
            doc_embs = doc_embs.reshape(1, -1)
        scores = [_cosine_sim(qe, doc_embs[i]) for i in range(len(candidates))]
        order = sorted(range(len(candidates)), key=lambda i: scores[i], reverse=True)
        return [candidates[i] for i in order]
    except Exception:
        return candidates


def _read_file(path: str) -> list[dict]:
    """CSV 또는 Excel 파일을 읽어 dict 리스트로 반환합니다. mtime 기반 캐시 사용."""
    global _file_cache
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = 0.0
    if path in _file_cache and _file_cache[path][0] == mtime:
        return _file_cache[path][1]

    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            # 첫 행 = 컬럼 제목. 넉넉히 읽어서 뒷열 누락 방지 (최소 30열)
            max_col = max(getattr(ws, "max_column", 0), 30)
            max_row = getattr(ws, "max_row", None) or 1
            rows = list(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True))
            if not rows:
                return []
            ncols = max_col
            first = list(rows[0])
            while len(first) < ncols:
                first.append(None)
            second = list(rows[1]) if len(rows) > 1 else None
            if second is not None:
                while len(second) < ncols:
                    second.append(None)
            headers = _normalize_headers(first, second)
            result = []
            for row in rows[1:]:
                row_list = list(row)
                while len(row_list) < ncols:
                    row_list.append(None)
                result.append({
                    headers[i]: (str(row_list[i]).strip() if row_list[i] is not None else "")
                    for i in range(ncols)
                })
            if len(_file_cache) >= CACHE_MAX_ENTRIES:
                _file_cache.pop(next(iter(_file_cache)))
            _file_cache[path] = (mtime, result)
            return result
        except ImportError:
            return []
    else:
        # 인코딩 자동 감지. 헤더 중복 시 _2, _3 ... 으로 고유화 후 읽기
        for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
            try:
                with open(path, "r", encoding=encoding) as f:
                    r = csv.reader(f)
                    first = next(r, None)
                if not first:
                    continue
                headers = _make_headers_unique([str(c).strip() or f"col{i}" for i, c in enumerate(first)])
                with open(path, "r", encoding=encoding) as f:
                    r = csv.reader(f)
                    next(r)
                    rows = []
                    for row in r:
                        row_list = list(row)
                        while len(row_list) < len(headers):
                            row_list.append("")
                        rows.append({headers[i]: (str(row_list[i]).strip() if i < len(row_list) else "") for i in range(len(headers))})
                if len(_file_cache) >= CACHE_MAX_ENTRIES:
                    _file_cache.pop(next(iter(_file_cache)))
                _file_cache[path] = (mtime, rows)
                return rows
            except (UnicodeDecodeError, UnicodeError):
                continue
        return []


def _read_headers(path: str) -> list[str] | None:
    """파일의 1행(헤더)만 읽어 컬럼명 리스트를 반환합니다. 빈 칸은 2행으로 보정."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            max_col = max(getattr(ws, "max_column", 0), 30)
            row1 = next(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col, values_only=True), None)
            if not row1:
                return None
            row1 = list(row1)
            while len(row1) < max_col:
                row1.append(None)
            row2_iter = ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=max_col, values_only=True)
            row2 = list(next(row2_iter, (None,) * max_col)) if max_col else None
            if row2 is not None:
                while len(row2) < max_col:
                    row2.append(None)
            return _normalize_headers(row1, row2)
        else:
            for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
                try:
                    with open(path, "r", encoding=enc) as f:
                        r = csv.reader(f)
                        first = next(r, None)
                    if not first:
                        return None
                    return _make_headers_unique([str(c).strip() or f"col{i}" for i, c in enumerate(first)])
                except (UnicodeDecodeError, UnicodeError):
                    continue
    except Exception:
        return None
    return None


def _detect_category_by_headers(headers: list[str]) -> str | None:
    """1행 컬럼명만으로 카테고리를 판별합니다. 헤더에 등장한 키워드 수로 매칭."""
    if not headers:
        return None
    header_text = " ".join(str(h).lower() for h in headers)
    best_label = None
    best_count = 0
    for label, keywords in HEADER_KEYWORDS.items():
        count = sum(1 for kw in keywords if kw.lower() in header_text)
        if count > best_count:
            best_count = count
            best_label = label
    return best_label if best_count > 0 else None


def _scan_directory() -> dict[str, str]:
    """
    WATCH_DIR 내 CSV/Excel 파일을 스캔하여
    { 카테고리: 파일경로 } 매핑을 반환합니다.
    같은 카테고리에 파일이 여러 개이면 가장 최근 수정 파일을 사용합니다.
    """
    candidates: dict[str, list[tuple[float, str]]] = {l: [] for l in LABELS}
    patterns = [
        os.path.join(WATCH_DIR, "*.csv"),
        os.path.join(WATCH_DIR, "**", "*.csv"),
    ]
    found_paths: set[str] = set()
    for pattern in patterns:
        for path in glob.glob(pattern, recursive=True):
            found_paths.add(os.path.abspath(path))

    for path in found_paths:
        fname = os.path.basename(path)
        if fname in IGNORE_FILENAMES:
            print(f"[파일 감지] '{fname}' → 스캔 제외")
            continue
        mtime = os.path.getmtime(path)
        headers = _read_headers(path)
        label = _detect_category_by_headers(headers) if headers else None
        if not label:
            label = "기타"
        candidates[label].append((mtime, path))
        print(f"[파일 감지] '{fname}' → {label}")

    def _pick_file(items):
        return sorted(items, key=lambda x: -x[0])[0][1]
    return {
        label: _pick_file(files)
        for label, files in candidates.items()
        if files
    }


def _get_file_index() -> dict[str, str]:
    """캐시된 파일 인덱스를 반환합니다. 만료됐으면 재스캔합니다."""
    global _file_index, _last_scan
    now = time.time()
    if SCAN_INTERVAL == 0 or (now - _last_scan) > SCAN_INTERVAL:
        _file_index = _scan_directory()
        _last_scan  = now
        print(f"[파일 인덱스] {_file_index}")
    return _file_index


def _filter_files_by_query_keywords(files_columns: list[dict], query: str) -> list[dict]:
    """
    질문에 나온 키워드가 컬럼에 포함된 파일만 남깁니다.
    예: '매출'이 들어 있으면 매출 컬럼이 있는 파일만 반환.
    """
    if not query or not files_columns:
        return files_columns
    # 2글자 이상 한글/영문/숫자 토큰 (숫자만 있는 건 제외)
    tokens = re.findall(r"[가-힣A-Za-z0-9_]{2,}", query)
    tokens = [t for t in tokens if not t.replace(".", "").replace("-", "").isdigit()]
    if not tokens:
        return files_columns
    columns_str_per_file = [" ".join(f.get("columns") or []) for f in files_columns]
    filtered = [
        f for f, col_str in zip(files_columns, columns_str_per_file)
        if any(t in col_str for t in tokens)
    ]
    return filtered if filtered else files_columns


def get_all_files_columns() -> list[dict]:
    """
    data 폴더(WATCH_DIR) 안의 모든 CSV/Excel 파일에 대해
    1행(헤더)만 스캔해 파일별 컬럼 목록을 반환합니다.
    반환: [ {"filename": "a.csv", "label": "재무", "columns": ["매출", "비용", ...]}, ... ]
    """
    patterns = [
        os.path.join(WATCH_DIR, "*.csv"),
        os.path.join(WATCH_DIR, "**", "*.csv"),
    ]
    found_paths: set[str] = set()
    for pattern in patterns:
        for path in glob.glob(pattern, recursive=True):
            found_paths.add(os.path.abspath(path))

    result: list[dict] = []
    for path in sorted(found_paths):
        fname = os.path.basename(path)
        if fname in IGNORE_FILENAMES:
            continue
        headers = _read_headers(path)
        if not headers:
            result.append({"filename": fname, "label": None, "columns": []})
            continue
        label = _detect_category_by_headers(headers) or "기타"
        result.append({"filename": fname, "label": label, "columns": list(headers)})
    return result


# ── 쿼리 핵심 로직 ─────────────────────────────────────────────────────────────

def _money_numeric_column_hint(q: str) -> str | None:
    """
    금액 비교 시 '어느 열'인지 질문에 명시된 경우에만 힌트.
    명시 없으면 None → 단가/매출/비용/예산 등 금액형 열을 두루 검사 (한 행에서 하나라도 조건 만족하면 통과).
    """
    if "단가" in q:
        return "단가"
    if "원가" in q:
        return "원가"
    if "매출" in q:
        return "매출"
    if "비용" in q:
        return "비용"
    if "예산" in q:
        return "예산"
    if "손익" in q:
        return "손익"
    if "미수금" in q:
        return "미수금"
    # 재고·제품 맥락: 금액 비교는 단가 열만 (재무표에서 매출·미수금 0이 'N만원 이하'에 전부 걸리는 오탐 방지)
    if any(k in q for k in ("제품", "품목", "부품", "자재")):
        if not any(
            k in q
            for k in ("매출", "비용", "예산", "손익", "미수금", "원가", "금액")
        ):
            return "단가"
    return None


def _normalize_query_for_numeric(query: str) -> str:
    """
    수치 조건 파싱 전 정규화.
    - NFKC (전각 숫자 등)
    - zero-width·BOM·NBSP 제거 후 공백 제거
    - 흔한 한글 금액 → 아라비아 숫자 표기
    """
    q = unicodedata.normalize("NFKC", (query or "").strip())
    q = re.sub(r"[\u200b-\u200d\ufeff\u00a0\u3000]", "", q)
    q = re.sub(r"\s+", "", q)
    # 긴 것부터 치환 (이십만원 → 십만원 오인 방지)
    _kr_man = [
        ("구십만원", "90만원"),
        ("팔십만원", "80만원"),
        ("칠십만원", "70만원"),
        ("육십만원", "60만원"),
        ("오십만원", "50만원"),
        ("사십만원", "40만원"),
        ("삼십만원", "30만원"),
        ("이십만원", "20만원"),
        ("십만원", "10만원"),
        ("백만원", "100만원"),
        ("이백만원", "200만원"),
        ("삼백만원", "300만원"),
        ("일만원", "1만원"),
        ("이만원", "2만원"),
        ("삼만원", "3만원"),
        ("사만원", "4만원"),
        ("오만원", "5만원"),
        ("육만원", "6만원"),
        ("칠만원", "7만원"),
        ("팔만원", "8만원"),
        ("구만원", "9만원"),
    ]
    for a, b in _kr_man:
        q = q.replace(a, b)
    return q


def _select_columns(query: str, columns: list[str], top_n: int = 0) -> list[str]:
    """조회 시 사용할 컬럼을 반환합니다. top_n=0이면 제한 없이 전부 반환합니다."""
    if top_n <= 0 or len(columns) <= top_n:
        return list(columns)
    # top_n이 양수일 때만 유사도 기반으로 일부만 선택 (과거 호환용)
    query_emb = _encode_query(query)
    col_embs  = _encode_docs(columns)
    sims      = [_cosine_sim(query_emb, e) for e in col_embs]
    ranked    = sorted(range(len(columns)), key=lambda i: sims[i], reverse=True)
    selected  = [columns[i] for i in ranked[:top_n]]
    return [c for c in columns if c in selected]


def _extract_search_tokens(query: str) -> list[str]:
    """
    질문에서 검색에 쓸 구체적인 토큰을 추출합니다.
    알파벳+숫자 혼합 토큰(M4x10, BLT-001 등)과 2글자 이상 한글 명사를 포함합니다.
    """
    import re
    stopwords = {"얼마야", "알려줘", "보여줘", "확인", "어떻게", "얼마나", "얼마",
                 "있어", "없어", "돼", "이번달", "오늘", "금일", "당일", "어제", "전일",
                 "내일", "익일", "이번주", "금주", "지난주", "전주", "다음주", "담주", "차주",
                 "지난달", "전월", "다음달", "차월", "올해", "금년", "작년", "전년", "내년", "익년",
                 "현재", "전체", "다",
                 "재고", "수량", "현황", "상태", "가격", "몇개", "몇 개",
                 "종류", "뭐", "어떤", "목록", "리스트", "전부", "모두", "뭐야",
                 "있나", "있어요", "있나요", "알려", "뭐있어", "뭐 있어",
                 "뭐가", "라는", "인데", "같은", "거야", "거예요", "인가요",
                 # 수치 비교 조사 — 행 데이터에 거의 없어 토큰 매칭만 망가뜨림
                 "이하", "이상", "초과", "미만", "이내", "보다", "넘는", "미만인", "이하인", "이상인"}
    tokens = re.findall(r"[A-Za-z0-9가-힣][\w가-힣x\-\.]*", query)
    filtered = []
    for t in tokens:
        if t in stopwords or len(t) < 2:
            continue
        # 수치 조건 토큰(예: 10만원, 1.5억원, 5만원대, 100개)은 텍스트 매칭에서 제외
        if re.fullmatch(r"\d+(?:\.\d+)?(?:억원|천만원|백만원|억|천만|백만|만원|만|천원|원|개|ea|EA)?(?:대)?", t):
            continue
        filtered.append(t)
    return filtered


def _parse_numeric_conditions(query: str, full_query: str | None = None) -> list[dict]:
    """
    질문에서 수치 비교 조건을 모두 파싱합니다.
    예) 10만원 이상, 1.5억 이하, 100개 미만, 5만원 이내, 이하 3만원, 2만원까지

    query: AND 조각 등 부분 문자열(숫자·단위는 여기서 추출).
    full_query: 원문 전체(공백 제거 정규화). 열 힌트(제품→단가 등)는 항상 여기 기준.
    """
    q = _normalize_query_for_numeric(query)
    qh = _normalize_query_for_numeric(full_query) if full_query is not None else q
    conditions = []
    unit_scale = {
        "억원": 100_000_000.0,
        "억": 100_000_000.0,
        "천만원": 10_000_000.0,
        "천만": 10_000_000.0,
        "백만원": 1_000_000.0,
        "백만": 1_000_000.0,
        "만원": 10_000.0,
        "만": 10_000.0,
        "천원": 1_000.0,
        "원": 1.0,
    }
    # 1) "N만원대/천원대/억대" 범위 조건
    range_pattern = r"(\d+(?:\.\d+)?)(억원|천만원|백만원|억|천만|백만|만원|만|천원|원)대"
    range_matches = re.findall(range_pattern, q)
    for num_str, unit in range_matches:
        base = float(num_str)
        step = unit_scale.get(unit, 1.0)
        low = base * step
        high = (base + 1.0) * step
        hint = _money_numeric_column_hint(qh)
        conditions.append({"kind": "money", "op": "range", "min": low, "max": high, "column_hint": hint})

    def _try_append_compare(num_str: str, unit: str, op_kw: str) -> None:
        """이상/초과/이하/미만/이내 한 건을 conditions에 넣습니다. 이내 → 이하와 동일(상한 포함)."""
        op_eff = "이하" if op_kw == "이내" else op_kw
        num = float(num_str)
        u = unit or ""
        column_hint = None
        if u in unit_scale:
            threshold = num * unit_scale[u]
            kind = "money"
            column_hint = _money_numeric_column_hint(qh)
        elif u in ("개", "ea", "EA"):
            threshold = num
            kind = "quantity"
        else:
            if any(
                k in qh
                for k in ["금액", "가격", "단가", "원가", "매출", "비용", "예산", "손익", "미수금"]
            ):
                threshold = num
                kind = "money"
                column_hint = _money_numeric_column_hint(qh)
            elif any(k in qh for k in ["수량", "재고", "개수", "물량", "qty", "quantity"]):
                threshold = num
                kind = "quantity"
            else:
                return
        conditions.append({"kind": kind, "op": op_eff, "threshold": threshold, "column_hint": column_hint})

    # 1-1) "만원대/억원대"를 1단위대와 동일하게 처리 (예: 만원대 == 1만원대)
    range_unit_only_pattern = r"(?<![\d\.])(억원|천만원|백만원|억|천만|백만|만원|만|천원|원)대"
    for unit in re.findall(range_unit_only_pattern, q):
        step = unit_scale.get(unit, 1.0)
        hint = _money_numeric_column_hint(qh)
        conditions.append({"kind": "money", "op": "range", "min": step, "max": step * 2.0, "column_hint": hint})

    # 2) "N만원 이상/이하/초과/미만/이내" 비교 조건
    compare_ops = r"이상|초과|이하|미만|이내"
    pattern = rf"(\d+(?:\.\d+)?)(억원|천만원|백만원|억|천만|백만|만원|만|원|개|ea|EA)?({compare_ops})"
    for num_str, unit, op_kw in re.findall(pattern, q):
        _try_append_compare(num_str, unit, op_kw)

    # 2-0b) "이하 3만원" / "이내 5만원" (역순)
    rev_pattern = rf"({compare_ops})(\d+(?:\.\d+)?)(억원|천만원|백만원|억|천만|백만|만원|만|천원|원|개|ea|EA)?"
    for op_kw, num_str, unit in re.findall(rev_pattern, q):
        _try_append_compare(num_str, unit, op_kw)

    # 2-0c) "3만원까지" → 이하(상한 포함)와 동일
    until_pattern = r"(\d+(?:\.\d+)?)(억원|천만원|백만원|억|천만|백만|만원|만|천원|원|개|ea|EA)?까지"
    for num_str, unit in re.findall(until_pattern, q):
        _try_append_compare(num_str, unit, "이하")

    # 2-1) "만원 이상"을 "1만원 이상"과 동일하게 처리
    # 단, "10만원이하"에서 만원 뒤의 '원'만 따로 잡혀 "1원 이하"가 붙는 오류 방지 → 단독 '원'은 만/천 직후가 아닐 때만
    unit_only_pattern = (
        rf"(?<![\d\.])(?:억원|천만원|백만원|억|천만|백만|만원|만|천원|(?<![만천])원)({compare_ops})"
    )
    for unit, op_kw in re.findall(unit_only_pattern, q):
        threshold = unit_scale.get(unit, 1.0)
        op_eff = "이하" if op_kw == "이내" else op_kw
        hint = _money_numeric_column_hint(qh)
        conditions.append({"kind": "money", "op": op_eff, "threshold": threshold, "column_hint": hint})
    return conditions


def _paren_covers_entire_string(s: str) -> bool:
    """s 전체가 한 쌍의 괄호로만 감싸졌는지 (중첩 포함)."""
    if len(s) < 2 or s[0] != "(" or s[-1] != ")":
        return False
    depth = 0
    for i, c in enumerate(s):
        if c == "(":
            depth += 1
        elif c == ")":
            depth -= 1
            if depth == 0:
                return i == len(s) - 1
    return False


def _strip_outer_parens_expr(s: str) -> str:
    s = s.strip()
    while _paren_covers_entire_string(s):
        s = s[1:-1].strip()
    return s


def _split_top_level_boolean(s: str, delimiters: tuple[str, ...]) -> list[str]:
    """괄호 밖에서만 delimiters로 분할. 없으면 [s] 한 덩어리."""
    delims = sorted(delimiters, key=len, reverse=True)
    depth = 0
    start = 0
    parts: list[str] = []
    i = 0
    n = len(s)
    while i < n:
        c = s[i]
        if c == "(":
            depth += 1
            i += 1
            continue
        if c == ")":
            depth = max(0, depth - 1)
            i += 1
            continue
        if depth == 0:
            matched = None
            for d in delims:
                if s.startswith(d, i):
                    matched = d
                    break
            if matched is not None:
                parts.append(s[start:i])
                start = i + len(matched)
                i = start
                continue
        i += 1
    parts.append(s[start:])
    out = [p.strip() for p in parts if p.strip()]
    return out if out else ([s.strip()] if s.strip() else [])


def parse_numeric_condition_groups(query: str) -> list[list[dict]]:
    """
    괄호 + 그리고(AND) + 또는(OR) 로 묶인 수치 조건을 파싱합니다.
    반환: OR 그룹들의 리스트, 각 그룹은 AND로 이어지는 조건 dict 목록.
    구분자 예: (단가5만이상그리고매출100만이상)또는재고50이상
    """
    q0 = _normalize_query_for_numeric(query)
    q0 = q0.replace("（", "(").replace("）", ")")
    if not q0:
        return []
    or_delims = ("||", "또는", "이거나", "OR", "or")
    and_delims = ("&&", "그리고", "이면서", "AND", "and")
    or_parts = _split_top_level_boolean(q0, or_delims)
    groups: list[list[dict]] = []
    for raw in or_parts:
        part = _strip_outer_parens_expr(raw)
        if not part:
            continue
        and_parts = _split_top_level_boolean(part, and_delims)
        conds: list[dict] = []
        for ap in and_parts:
            ap = _strip_outer_parens_expr(ap)
            if ap:
                conds.extend(_parse_numeric_conditions(ap, full_query=q0))
        if conds:
            groups.append(conds)
    return groups


def _flatten_numeric_groups(groups: list[list[dict]]) -> list[dict]:
    return [c for g in groups for c in g]


def _numeric_value_columns(
    columns: list[str], kind: str, column_hint: str | None = None, query: str | None = None
) -> list[str]:
    """
    조건 종류(kind)에 맞는 수치 컬럼 후보를 반환합니다.
    column_hint 또는 query에서 뽑은 토큰이 컬럼명에 있으면, 그 열에서만 조회합니다.
    """
    if kind == "money":
        keys = [
            "금액",
            "가격",
            "단가",
            "원가",
            "매출",
            "비용",
            "예산",
            "손익",
            "미수금",
            "price",
            "amount",
        ]
    else:
        keys = ["수량", "재고", "개수", "물량", "qty", "quantity"]
    lowered_keys = [k.lower() for k in keys]
    kind_cols = [c for c in columns if any(k in c.lower() for k in lowered_keys)]

    # 쿼리에 나온 단어가 컬럼명에 있으면 그 열만 사용
    if column_hint:
        hint_lower = column_hint.lower()
        pri = [c for c in kind_cols if hint_lower in c.lower()]
        return pri if pri else []
    if query:
        tokens = _extract_search_tokens(query)
        if tokens:
            pri = [c for c in kind_cols if any(t.lower() in c.lower() for t in tokens)]
            if pri:
                return pri
    return kind_cols if kind_cols else columns


def _resolve_numeric_columns_for_condition(
    columns: list[str], cond: dict, query: str | None
) -> list[str]:
    """
    한 건의 수치 조건에 대해 비교할 컬럼 순서를 정합니다.
    column_hint·토큰이 있으면 기존 규칙, 없고 후보가 여럿이면 질문–컬럼명 임베딩 순으로 정렬합니다.
    """
    hint = cond.get("column_hint")
    base = _numeric_value_columns(columns, cond["kind"], column_hint=hint, query=query)
    if not base:
        # 힌트가 있는데 이 파일에 그 열이 없으면 '전 금액 열'로 넓히지 않음
        # (예: 단가 힌트인데 finance.csv만 있으면 매출 0이 N만원 이하에 전부 걸림)
        if hint:
            return []
        base = _numeric_value_columns(columns, cond["kind"], column_hint=None, query=None)
    if not base:
        return []
    if len(base) == 1:
        return base
    if hint:
        return base
    return _embed_rank_numeric_columns(base, query or "")


def _to_float_value(raw) -> float | None:
    text = str(raw or "").strip()
    if not text:
        return None
    cleaned = (
        text.replace(",", "")
        .replace("원", "")
        .replace("개", "")
        .replace("ea", "")
        .replace("EA", "")
        .replace("%", "")
        .strip()
    )
    try:
        return float(cleaned)
    except ValueError:
        return None


def _match_numeric(value: float, op: str, threshold: float) -> bool:
    if op == "이상":
        return value >= threshold
    if op == "초과":
        return value > threshold
    if op == "이하":
        return value <= threshold
    if op == "미만":
        return value < threshold
    return False


def _row_satisfies_one_numeric_cond(row: dict, columns: list[str], cond: dict, query: str | None) -> bool:
    candidate_cols = _resolve_numeric_columns_for_condition(columns, cond, query)
    for col in candidate_cols:
        value = _to_float_value(row.get(col))
        if value is None:
            continue
        if cond["op"] == "range":
            low = float(cond["min"])
            high = float(cond["max"])
            if low <= value < high:
                return True
        elif _match_numeric(value, cond["op"], float(cond["threshold"])):
            return True
    return False


def _row_satisfies_all_numeric_conds(row: dict, columns: list[str], conditions: list[dict], query: str | None) -> bool:
    if not conditions:
        return True
    return all(_row_satisfies_one_numeric_cond(row, columns, c, query) for c in conditions)


def _apply_numeric_conditions(
    rows: list[dict], columns: list[str], conditions: list[dict], query: str | None = None
) -> list[dict]:
    """AND: conditions 전부 만족하는 행만 유지."""
    if not conditions or not rows:
        return rows
    return [row for row in rows if _row_satisfies_all_numeric_conds(row, columns, conditions, query)]


def _apply_numeric_condition_groups(
    rows: list[dict], columns: list[str], groups: list[list[dict]], query: str | None = None
) -> list[dict]:
    """
    OR 그룹: 그룹 하나라도 AND 조건을 모두 만족하면 행 통과.
    그룹이 하나뿐이면 기존과 동일하게 AND만 적용.
    """
    if not rows:
        return rows
    flat = [g for g in groups if g]
    if not flat:
        return rows
    if len(flat) == 1:
        return _apply_numeric_conditions(rows, columns, flat[0], query=query)
    return [row for row in rows if any(_row_satisfies_all_numeric_conds(row, columns, g, query) for g in flat)]


def _contains_any(q: str, aliases: tuple[str, ...]) -> bool:
    return any(a in q for a in aliases)


def _extract_time_window(query: str) -> tuple[date, date, str] | None:
    today = date.today()
    q = query.replace(" ", "")
    today_aliases = ("오늘", "금일", "당일")
    yesterday_aliases = ("어제", "전일")
    tomorrow_aliases = ("내일", "익일", "명일")
    this_week_aliases = ("이번주", "금주", "당주", "이번주간")
    last_week_aliases = ("지난주", "저번주", "전주", "직전주")
    next_week_aliases = ("다음주", "담주", "차주", "익주")
    this_month_aliases = ("이번달", "금월", "당월")
    last_month_aliases = ("지난달", "저번달", "전월", "직전월")
    next_month_aliases = ("다음달", "차월", "익월")
    this_year_aliases = ("올해", "금년", "당해")
    last_year_aliases = ("작년", "전년")
    next_year_aliases = ("내년", "익년", "명년")

    if _contains_any(q, today_aliases):
        return today, today, "오늘"
    if _contains_any(q, yesterday_aliases):
        y = today - timedelta(days=1)
        return y, y, "어제"
    if _contains_any(q, tomorrow_aliases):
        t = today + timedelta(days=1)
        return t, t, "내일"

    if _contains_any(q, this_week_aliases):
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start, end, "이번주"
    if _contains_any(q, last_week_aliases):
        this_week_start = today - timedelta(days=today.weekday())
        start = this_week_start - timedelta(days=7)
        end = this_week_start - timedelta(days=1)
        return start, end, "지난주"
    if _contains_any(q, next_week_aliases):
        this_week_start = today - timedelta(days=today.weekday())
        start = this_week_start + timedelta(days=7)
        end = start + timedelta(days=6)
        return start, end, "다음주"

    if _contains_any(q, this_month_aliases):
        start = today.replace(day=1)
        if start.month == 12:
            next_month = start.replace(year=start.year + 1, month=1, day=1)
        else:
            next_month = start.replace(month=start.month + 1, day=1)
        end = next_month - timedelta(days=1)
        return start, end, "이번달"

    if _contains_any(q, last_month_aliases):
        this_month_start = today.replace(day=1)
        end = this_month_start - timedelta(days=1)
        start = end.replace(day=1)
        return start, end, "지난달"
    if _contains_any(q, next_month_aliases):
        this_month_start = today.replace(day=1)
        if this_month_start.month == 12:
            next_month_start = this_month_start.replace(year=this_month_start.year + 1, month=1, day=1)
        else:
            next_month_start = this_month_start.replace(month=this_month_start.month + 1, day=1)
        if next_month_start.month == 12:
            after_next_month_start = next_month_start.replace(year=next_month_start.year + 1, month=1, day=1)
        else:
            after_next_month_start = next_month_start.replace(month=next_month_start.month + 1, day=1)
        end = after_next_month_start - timedelta(days=1)
        return next_month_start, end, "다음달"

    if _contains_any(q, this_year_aliases):
        start = date(today.year, 1, 1)
        end = date(today.year, 12, 31)
        return start, end, "올해"
    if _contains_any(q, last_year_aliases):
        y = today.year - 1
        return date(y, 1, 1), date(y, 12, 31), "작년"
    if _contains_any(q, next_year_aliases):
        y = today.year + 1
        return date(y, 1, 1), date(y, 12, 31), "내년"

    return None


def _detect_flow_intent(query: str) -> str | None:
    q = (query or "").replace(" ", "")
    wants_in = any(k in q for k in ["입고", "입하"])
    wants_out = any(k in q for k in ["출고", "출하"])
    wants_return = any(k in q for k in ["반품", "리턴", "환입"])
    if wants_return:
        return "return"
    if wants_in and not wants_out:
        return "in"
    if wants_out and not wants_in:
        return "out"
    return None


def _infer_flow_columns(rows: list[dict], columns: list[str]) -> dict[str, list[str]]:
    """
    입고/출고/반품 관련 컬럼을 자동 추론합니다.
    - 1차: 컬럼명 키워드 기반
    - 2차: 샘플값(텍스트) 키워드 기반
    """
    keyword_map = {
        "in": ["입고", "입하"],
        "out": ["출고", "출하"],
        "return": ["반품", "리턴", "환입"],
    }
    inferred: dict[str, set[str]] = {k: set() for k in keyword_map}

    # 1) 컬럼명 기반 추론
    for col in columns:
        col_text = str(col)
        for flow, keys in keyword_map.items():
            if any(k in col_text for k in keys):
                inferred[flow].add(col)

    # 2) 샘플값 기반 추론 (컬럼명 단서가 약한 파일 대응)
    sample_rows = rows[:200]
    hit_counts: dict[str, dict[str, int]] = {flow: {c: 0 for c in columns} for flow in keyword_map}
    for row in sample_rows:
        for col in columns:
            val = str(row.get(col, "")).strip()
            if not val:
                continue
            for flow, keys in keyword_map.items():
                if any(k in val for k in keys):
                    hit_counts[flow][col] += 1

    for flow in keyword_map:
        if inferred[flow]:
            continue
        for col in columns:
            if hit_counts[flow].get(col, 0) >= 2:
                inferred[flow].add(col)

    return {flow: list(cols) for flow, cols in inferred.items()}


def _apply_flow_filter(query: str, rows: list[dict], columns: list[str]) -> list[dict]:
    """
    입고/출고 질의 시 반대 방향 데이터가 섞이지 않도록 필터합니다.
    - 출고 질의인데 출고 컬럼이 없고 입고 컬럼만 있으면 빈 결과 반환
    - 입고 질의도 동일하게 반대 케이스 처리
    """
    flow = _detect_flow_intent(query)
    if not flow or not rows:
        return rows

    keyword_map = {
        "in": ["입고", "입하"],
        "out": ["출고", "출하"],
        "return": ["반품", "리턴", "환입"],
    }
    target_keys = keyword_map.get(flow, [])
    flow_cols = _infer_flow_columns(rows, columns)

    target_cols = flow_cols.get(flow, [])
    opposite_cols = []
    for other_flow, cols in flow_cols.items():
        if other_flow != flow:
            opposite_cols.extend(cols)

    # 컬럼에 직접 없더라도 값 텍스트에서 타겟 키워드가 발견되면 해당 행만 유지
    if not target_cols:
        text_matched = []
        for row in rows:
            row_text = " ".join(str(v) for v in row.values())
            if any(k in row_text for k in target_keys):
                text_matched.append(row)
        if text_matched:
            return text_matched
        if opposite_cols:
            return []
        # 방향 단서를 전혀 추론할 수 없으면 필터를 강제하지 않습니다.
        return rows

    filtered = []
    for row in rows:
        has_target = False
        for col in target_cols:
            val = str(row.get(col, "")).strip()
            if val and val.lower() not in {"nan", "none", "null", "-"}:
                has_target = True
                break
        if has_target:
            filtered.append(row)
    return filtered


def _is_date_column(col: str) -> bool:
    keywords = [
        "일자",
        "날짜",
        "date",
        "일시",
        "생성일",
        "등록일",
        "작성일",
        "기준일",
        "예정일",
        "입고일",
        "출고일",
        "납기일",
    ]
    cl = col.lower()
    return any(k in col for k in keywords) or any(k in cl for k in keywords)


def _parse_date_value(raw) -> date | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None

    # 시간 정보가 포함되면 날짜 부분만 우선 파싱
    text = text.split(" ")[0]
    text = text.split("T")[0]
    text = text.replace(" ", "")

    fmts = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%Y%m%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d.%m.%Y",
        "%d/%m/%y",
        "%d-%m-%y",
        "%d.%m.%y",
        "%y-%m-%d",
        "%y/%m/%d",
        "%y.%m.%d",
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%m.%d.%Y",
        "%m/%d/%y",
        "%m-%d-%y",
        "%m.%d.%y",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    # 연도 없는 표현(예: 22/04)은 현재 연도로 보정 시도
    for short_fmt in ("%d/%m", "%d-%m", "%d.%m", "%m/%d", "%m-%d", "%m.%d"):
        try:
            dt = datetime.strptime(text, short_fmt)
            return date(date.today().year, dt.month, dt.day)
        except ValueError:
            continue
    return None


# 주 단위 요청 시 같은 달 데이터도 포함할 때 쓰는 구간 이름 (월 단위 데이터 보정용)
_WEEK_WINDOW_NAMES = ("이번주", "지난주", "다음주", "금주", "당주", "저번주", "전주", "담주", "차주")
# 하루 단위 요청 시 같은 달 데이터도 포함 (어제/오늘/내일 등)
_DAY_WINDOW_NAMES = ("오늘", "금일", "당일", "어제", "전일", "내일", "익일", "명일")
# 연 단위 요청 시 같은 해 데이터도 포함 (올해/작년/내년 등)
_YEAR_WINDOW_NAMES = ("올해", "금년", "당해", "작년", "전년", "내년", "익년", "명년")


def _embedding_text_columns(columns: list[str], rows: list[dict]) -> list[str]:
    """
    _filter_rows 2단계 임베딩에 쓰는 텍스트 컬럼 목록 (기존 로직과 동일).
    사전 임베딩 캐시는 이 목록이 파일 전체 기준으로 동일할 때만 사용합니다.
    """
    return [
        c
        for c in columns
        if any(
            not str(row.get(c, "")).replace(".", "").replace("-", "").replace("%", "").isdigit()
            and len(str(row.get(c, ""))) > 1
            for row in rows[:5]
        )
    ]


def _get_stored_full_row_embeddings(source_path: str, text_cols: list[str]) -> np.ndarray | None:
    """디스크/메모리 캐시에서 소스 파일 전체 행 임베딩 (n, dim) 또는 None."""
    if not _USE_ROW_EMB_CACHE or not source_path:
        return None
    abs_p = os.path.abspath(source_path)
    try:
        mtime = os.path.getmtime(abs_p)
    except OSError:
        return None
    tc = tuple(text_cols)
    mem_key = (abs_p, mtime, tc, _EMBEDDING_MODEL_ID)
    if mem_key in _FULL_ROW_EMB_MEM:
        return _FULL_ROW_EMB_MEM[mem_key]
    emb = load_row_embedding_cache(abs_p, list(text_cols), _EMBEDDING_MODEL_ID)
    if emb is not None:
        _FULL_ROW_EMB_MEM[mem_key] = emb
    return emb


def _filter_rows_by_time(query: str, rows: list[dict], all_columns: list[str]) -> tuple[list[dict], str | None, bool]:
    window = _extract_time_window(query)
    if not window or not rows:
        return rows, None, False

    start_date, end_date, window_name = window
    date_cols = [c for c in all_columns if _is_date_column(c)]
    if not date_cols:
        return [], f"기간 필터 요청: {window_name} (날짜 컬럼 없음)", True

    # 주/일 단위 요청이면 같은 (년,월) 행 포함, 연 단위면 같은 연도 행 포함 (월/연 단위 데이터 보정용)
    allow_same_month = window_name in _WEEK_WINDOW_NAMES or window_name in _DAY_WINDOW_NAMES
    allow_same_year = window_name in _YEAR_WINDOW_NAMES

    filtered = []
    for row in rows:
        parsed_any = False
        for col in date_cols:
            d = _parse_date_value(row.get(col))
            if d is None:
                continue
            parsed_any = True
            if start_date <= d <= end_date:
                filtered.append(row)
                break
            # 이번주/오늘 등: 데이터가 월 단위(말일 등)면 같은 달이면 포함
            if allow_same_month and (d.year == start_date.year and d.month == start_date.month):
                filtered.append(row)
                break
            # 올해/작년 등: 같은 해 행 포함 (연 단위 데이터 보정)
            if allow_same_year and d.year == start_date.year:
                filtered.append(row)
                break
        if not parsed_any:
            continue

    note = f"기간 필터: {window_name} ({start_date}~{end_date})"
    if allow_same_month and filtered:
        note = f"기간 필터: {window_name} ({start_date}~{end_date}, 동일 월 포함)"
    elif allow_same_year and filtered:
        note = f"기간 필터: {window_name} ({start_date}~{end_date}, 동일 연 포함)"
    return filtered, note, False


def _filter_rows(
    query: str,
    rows: list[dict],
    columns: list[str],
    *,
    source_path: str | None = None,
    full_row_index: dict[int, int] | None = None,
) -> list[dict]:
    if not rows:
        return rows

    rows = _apply_flow_filter(query, rows, columns)
    if not rows:
        return []

    numeric_groups = parse_numeric_condition_groups(query)
    rows = _apply_numeric_condition_groups(rows, columns, numeric_groups, query=query)
    if not rows:
        return []

    has_numeric = any(bool(g) for g in numeric_groups)

    if _USE_KEYWORD_ROW_FILTER:
        tokens = _extract_search_tokens(query)
        if has_numeric:
            tokens = [
                t
                for t in tokens
                if t not in _NUMERIC_QUERY_SKIP_TOKENS and t.lower() not in _NUMERIC_QUERY_SKIP_TOKENS
            ]

        # 수치만 있고 토큰 없으면 임베딩 생략(구 방식)
        if has_numeric and not tokens:
            return rows

        def _token_in_row(t: str, row_text_lower: str) -> bool:
            """토큰 전체 또는 긴 토큰(4글자 이상)의 2글자 이상 부분 문자열이 행에 있으면 True."""
            tl = t.lower()
            if tl in row_text_lower:
                return True
            if len(tl) >= 4:
                for n in range(2, len(tl)):
                    for i in range(len(tl) - n + 1):
                        if tl[i : i + n] in row_text_lower:
                            return True
            return False

        if tokens:
            exact_matched = []
            for row in rows:
                col_str = " ".join(str(c) for c in columns).lower()
                val_str = " ".join(str(v) for v in row.values()).lower()
                row_text = f"{col_str} {val_str}"
                if all(_token_in_row(t, row_text) for t in tokens):
                    exact_matched.append(row)
            if not exact_matched:
                for row in rows:
                    col_str = " ".join(str(c) for c in columns).lower()
                    val_str = " ".join(str(v) for v in row.values()).lower()
                    row_text = f"{col_str} {val_str}"
                    if any(_token_in_row(t, row_text) for t in tokens):
                        exact_matched.append(row)
            if exact_matched:
                return exact_matched
            if has_numeric and rows:
                return rows
            return []

    # ── 임베딩 유사도 검색 (기본: 키워드 단계 없이 항상) ──────────────────
    text_cols = _embedding_text_columns(columns, rows)
    if not text_cols:
        return rows

    # 단가(금액) 조건만 있고, 검색 토큰이 제품/품목 같은 일반어뿐이면 임베딩 순서가 오히려 방해 → 단가순 전체(또는 TOP_K)
    flat_money = _flatten_numeric_groups(numeric_groups)
    only_money_conds = bool(flat_money) and all(c.get("kind") == "money" for c in flat_money)
    if only_money_conds and has_numeric:
        spec_toks = _extract_search_tokens(query)
        spec_toks = [
            t
            for t in spec_toks
            if t not in _NUMERIC_QUERY_SKIP_TOKENS and t.lower() not in _NUMERIC_QUERY_SKIP_TOKENS
        ]
        if not spec_toks:
            price_col = next((c for c in columns if "단가" in str(c)), None)
            if price_col:

                def _row_unit_price(r: dict) -> float:
                    v = _to_float_value(r.get(price_col))
                    return float("inf") if v is None else v

                want_high_first = any(c.get("op") in ("이상", "초과") for c in flat_money)
                out = sorted(rows, key=_row_unit_price, reverse=want_high_first)
                if _ROW_EMBED_TOP_K > 0:
                    out = out[: _ROW_EMBED_TOP_K]
                return out

    query_emb = _encode_query(query)
    row_embs: np.ndarray | None = None
    if source_path and full_row_index is not None:
        full_stored = _get_stored_full_row_embeddings(source_path, text_cols)
        if full_stored is not None:
            try:
                idxs = [full_row_index[id(row)] for row in rows]
                if idxs and all(0 <= i < full_stored.shape[0] for i in idxs):
                    row_embs = full_stored[idxs]
            except KeyError:
                row_embs = None
    if row_embs is None:
        row_texts = [" ".join(str(row.get(c, "")) for c in text_cols) for row in rows]
        row_embs = _encode_docs(row_texts)

    sims = [_cosine_sim(query_emb, row_embs[i]) for i in range(len(rows))]
    pairs: list[tuple[dict, float]] = list(zip(rows, sims))
    pairs.sort(key=lambda x: -x[1])

    thr = _ROW_EMBED_SIM_THRESHOLD
    if thr is not None:
        kept = [(r, s) for r, s in pairs if s >= thr]
        out = [r for r, _ in kept] if kept else list(rows)
    else:
        out = [r for r, _ in pairs]

    if _ROW_EMBED_TOP_K > 0:
        out = out[: _ROW_EMBED_TOP_K]

    return out


def _trim_trailing_empty_columns(columns: list[str], rows: list[dict]) -> tuple[list[str], list[dict]]:
    """
    끝쪽의 비어 있는 컬럼(col N 또는 실제로 값이 없는 컬럼)을 제거해 표시용 컬럼 수를 맞춥니다.
    """
    if not columns or not rows:
        return columns, rows
    # 마지막으로 의미 있는 컬럼 인덱스: 컬럼명이 col숫자이면서 모든 행에서 빈 값이면 제외
    last_keep = -1
    for i in range(len(columns)):
        col = columns[i]
        is_generic = bool(re.match(r"^col\d+$", col))
        has_any = any(str(row.get(col, "")).strip() for row in rows)
        if not is_generic or has_any:
            last_keep = i
    if last_keep < 0:
        return columns[:1], [{c: row.get(c, "") for c in columns[:1]} for row in rows]
    keep_cols = columns[: last_keep + 1]
    keep_rows = [{c: row.get(c, "") for c in keep_cols} for row in rows]
    return keep_cols, keep_rows


def _make_summary(rows: list[dict], columns: list[str]) -> str:
    total = len(rows)
    numeric_stats = []
    for col in columns:
        values = []
        for row in rows:
            try:
                values.append(float(str(row.get(col, "")).replace(",", "").replace("%", "")))
            except ValueError:
                pass
        if values and any(kw in col for kw in ["수량", "재고", "생산", "금액", "비용", "매출", "달성", "가동", "이익", "손익", "순이익"]):
            numeric_stats.append(f"{col} 합계: {sum(values):,.0f}")
    base = f"총 {total}개 항목 조회"
    return base + " | " + " | ".join(numeric_stats[:3]) if numeric_stats else base


# ── 공개 API ──────────────────────────────────────────────────────────────────

def _get_all_data_paths() -> list[str]:
    """WATCH_DIR 안의 조회 대상 CSV 파일 경로 목록을 반환합니다 (정렬, 제외 파일 제외). xlsx/xls는 스캔하지 않음."""
    patterns = [
        os.path.join(WATCH_DIR, "*.csv"),
        os.path.join(WATCH_DIR, "**", "*.csv"),
    ]
    found: set[str] = set()
    for pattern in patterns:
        for path in glob.glob(pattern, recursive=True):
            if os.path.basename(path) not in IGNORE_FILENAMES:
                found.add(os.path.abspath(path))
    return sorted(found)


def _query_single_file(path: str, user_query: str, use_time_filter: bool = True) -> dict | None:
    """
    한 파일에 대해 조회를 수행합니다. 조건에 맞는 행이 있으면 결과 dict, 없으면 None.
    use_time_filter=False면 기간 필터를 쓰지 않고 전체 행을 대상으로 합니다.
    """
    all_rows = _read_file(path)
    if not all_rows:
        return None
    all_columns = list(all_rows[0].keys())
    row_id_to_index = {id(r): i for i, r in enumerate(all_rows)}
    if use_time_filter:
        time_filtered_rows, time_filter_note, time_filter_blocked = _filter_rows_by_time(
            user_query, all_rows, all_columns
        )
    else:
        time_filtered_rows = all_rows
        time_filter_note = "해당 기간 데이터 없음 → 전체 기간 표시"
        time_filter_blocked = False

    selected_cols = _select_columns(user_query, all_columns)
    numeric_groups_q = parse_numeric_condition_groups(user_query)
    flat_numeric = _flatten_numeric_groups(numeric_groups_q)
    if flat_numeric:
        for cond in flat_numeric:
            for col in _resolve_numeric_columns_for_condition(all_columns, cond, user_query)[:2]:
                if col not in selected_cols:
                    selected_cols.append(col)
    filtered_rows = _filter_rows(
        user_query,
        time_filtered_rows,
        all_columns,
        source_path=path,
        full_row_index=row_id_to_index,
    )
    trimmed_rows = [{c: row.get(c, "") for c in selected_cols} for row in filtered_rows]
    display_cols, display_rows = _trim_trailing_empty_columns(selected_cols, trimmed_rows)
    if not display_rows:
        return None

    if time_filter_blocked:
        summary = (
            f"{time_filter_note} | 날짜 기반 조회를 수행할 수 없습니다. "
            "데이터 파일에 날짜/일자/예정일 컬럼을 추가해 주세요."
        )
    else:
        summary = _make_summary(display_rows, display_cols)
        if time_filter_note:
            summary = f"{time_filter_note} | {summary}"

    used_filename = os.path.basename(path)
    label = _detect_category_by_headers(all_columns) or "기타"
    used_files_columns = [{"filename": used_filename, "label": label, "columns": display_cols}]
    return {
        "label":             label,
        "filename":          used_filename,
        "summary":           summary,
        "columns":           display_cols,
        "rows":              [list(r.values()) for r in display_rows],
        "time_filter_blocked": time_filter_blocked,
        "all_files_columns": used_files_columns,
    }


# 파일 우선순위 계산 시 제외할 기간/일자 단어 (기간 단어로는 파일 선택 우선순위를 주지 않음)
TIME_PERIOD_TOKENS = {
    "오늘", "금일", "당일", "어제", "전일", "내일", "익일", "명일",
    "이번주", "금주", "당주", "이번주간", "지난주", "저번주", "전주", "직전주",
    "다음주", "담주", "차주", "익주",
    "이번달", "금월", "당월", "지난달", "저번달", "전월", "직전월",
    "다음달", "차월", "익월",
    "올해", "금년", "당해", "작년", "전년", "내년", "익년", "명년",
    "현재", "전체", "기간", "일자", "날짜",
}


def _query_tokens(query: str) -> list[str]:
    """질문에서 컬럼 매칭용 키워드 토큰을 뽑습니다."""
    if not query:
        return []
    tokens = re.findall(r"[가-힣A-Za-z0-9_]{2,}", query)
    return [t for t in tokens if not t.replace(".", "").replace("-", "").isdigit()]


def _paths_sorted_by_relevance(paths: list[str], user_query: str) -> tuple[list[str], list[str]]:
    """
    질문 키워드(기간 제외)가 컬럼에 맞는 파일만 우선 정렬해 반환합니다.
    반환: (키워드 맞는 파일들 정렬, 그 외 파일들). 키워드가 없으면 (전체 정렬, []).
    """
    tokens = [t for t in _query_tokens(user_query) if t not in TIME_PERIOD_TOKENS]
    all_paths_ordered = list(paths)
    all_files = get_all_files_columns()
    by_fname = {f["filename"]: " ".join(f.get("columns") or []) for f in all_files}

    qn = re.sub(r"\s+", "", (user_query or "").strip())
    qn_norm = _normalize_query_for_numeric(user_query or "")
    # "십만원 이하"처럼 한글만 있어도 정규화 후 숫자가 생김 → 재고 CSV(단가 열) 우선
    # "10만원 이하 제품" 처럼 토큰이 컬럼명과 안 겹쳐도, 수치 비교 질문이면 단가/금액 열 있는 파일을 앞으로
    money_compare_q = bool(re.search(r"\d", qn_norm)) and any(
        k in qn_norm
        for k in ("이하", "이상", "초과", "미만", "이내", "까지", "만원", "억", "천만", "백만", "원")
    )

    def score(path: str) -> int:
        col_str = by_fname.get(os.path.basename(path), "")
        col_lower = col_str.lower()
        s = sum(1 for t in tokens if t in col_str)
        if money_compare_q:
            for h in ("단가", "가격", "매출", "비용", "원가", "price", "amount", "unit"):
                if h.lower() in col_lower:
                    s += 3
                    break
        return s

    with_score = [(path, score(path)) for path in paths]
    relevant_sorted = [path for path, s in sorted(with_score, key=lambda x: x[1], reverse=True) if s >= 1]
    others = [path for path, s in with_score if s == 0]
    if not tokens and not money_compare_q:
        return (all_paths_ordered, [])
    if not tokens and money_compare_q:
        return (relevant_sorted, others) if relevant_sorted else (all_paths_ordered, [])
    return (relevant_sorted, others)


def _merge_multi_file_results(path_results: list[tuple[str, dict]]) -> dict:
    """
    여러 파일 조회 결과를 하나로 합칩니다.
    컬럼은 합집합(중복 이름은 하나로), 각 행 앞에 '파일' 컬럼을 붙입니다.
    """
    if not path_results:
        return None
    if len(path_results) == 1:
        path, one = path_results[0]
        cols = one.get("columns") or []
        rows = one.get("rows") or []
        return {
            "label":       one.get("label", "기타"),
            "filename":    os.path.basename(path),
            "summary":     one.get("summary", ""),
            "columns":    ["파일"] + cols,
            "rows":       [[os.path.basename(path)] + r for r in rows],
            "time_filter_blocked": one.get("time_filter_blocked", False),
            "all_files_columns": one.get("all_files_columns", []),
        }

    all_columns_set: set[str] = set()
    for _, res in path_results:
        for c in res.get("columns") or []:
            all_columns_set.add(c)
    merged_cols = ["파일"] + sorted(all_columns_set)
    merged_rows: list[list] = []
    summaries: list[str] = []

    for path, res in path_results:
        fname = os.path.basename(path)
        cols = res.get("columns") or []
        rows = res.get("rows") or []
        for r in rows:
            row_list = [fname]
            for c in merged_cols[1:]:
                if c in cols:
                    idx = cols.index(c)
                    row_list.append(r[idx] if idx < len(r) else "")
                else:
                    row_list.append("")
            merged_rows.append(row_list)
        if res.get("summary"):
            summaries.append(res["summary"])

    combined_summary = " | ".join(summaries[:3]) if summaries else f"총 {len(merged_rows)}개 항목 (여러 파일)"
    used_files = [
        {"filename": os.path.basename(p), "label": r.get("label", "기타"), "columns": r.get("columns") or []}
        for p, r in path_results
    ]
    return {
        "label":             "기타",
        "filename":          None,
        "summary":           combined_summary,
        "columns":           merged_cols,
        "rows":              merged_rows,
        "time_filter_blocked": any(r.get("time_filter_blocked") for _, r in path_results),
        "all_files_columns": used_files,
    }


def query_data(label: str | None, user_query: str) -> dict | None:
    """
    매 질문마다 모든 파일을 돌아가며 조회하고,
    조건에 맞는 행이 나온 파일들은 한꺼번에 합쳐서 조회 데이터로 보여줍니다.
    (중복·비슷한 컬럼이 있으면 하나의 테이블로 합침)
    """
    paths = _get_all_data_paths()
    keyword_matched, rest = _paths_sorted_by_relevance(paths, user_query)
    to_try = keyword_matched + rest

    path_results: list[tuple[str, dict]] = []
    for path in to_try:
        result = _query_single_file(path, user_query)
        if result and result.get("rows"):
            path_results.append((path, result))

    # 모든 파일을 훑어서 검색어가 행 내용에 포함된 결과는 전부 사용 (컬럼명에만 있는 파일로 한정하지 않음)
    if not path_results:
        # 기간 필터로 걸린 경우 기간 없이 재시도
        for path in to_try:
            result = _query_single_file(path, user_query, use_time_filter=False)
            if result and result.get("rows"):
                path_results.append((path, result))

    if not path_results:
        all_files_columns = get_all_files_columns()
        shown_files = _filter_files_by_query_keywords(all_files_columns, user_query)
        return {
            "label":             "기타",
            "filename":          None,
            "summary":           "조건에 맞는 데이터가 없습니다.",
            "columns":           [],
            "rows":              [],
            "all_files_columns": shown_files,
        }

    return _merge_multi_file_results(path_results)


def get_watch_dir() -> str:
    return WATCH_DIR


def rescan() -> dict[str, str]:
    """강제로 파일 인덱스를 재스캔하고 파일 내용 캐시를 비웁니다."""
    global _last_scan, _file_cache, _FULL_ROW_EMB_MEM
    _last_scan = 0.0
    _file_cache.clear()
    _FULL_ROW_EMB_MEM.clear()
    return _get_file_index()
